"""End-to-end regression test: build the fixture study and confirm the comparison finds
exactly the differences the fixture plants, and no others.

    python -m pytest sdtm_builder/tests/test_smoke.py       (or run this file directly)
"""
from __future__ import annotations

import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sdtm_builder.build import build_study                       # noqa: E402
from sdtm_builder.compare import compare_study                   # noqa: E402
from sdtm_builder.rawio import RawStore                          # noqa: E402
from sdtm_builder.spec import load_spec                          # noqa: E402

HERE = Path(__file__).resolve().parent


def _fixture(tmp: Path) -> Path:
    subprocess.run([sys.executable, str(HERE / "make_fixture.py"), str(tmp)],
                   check=True, capture_output=True)
    return tmp


def _load(tmp: Path):
    spec = load_spec(tmp / "mapping_spec.xlsx")
    store = RawStore.discover(tmp / "raw")
    return spec, store, build_study(spec, store)


def test_build_shapes():
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        _spec, _store, results = _load(tmp)
        assert set(results) == {"DM", "AE", "VS", "EG", "DS"}
        assert all(r.ok for r in results.values())
        assert len(results["DM"].dataset) == 6
        assert len(results["AE"].dataset) == 8
        assert len(results["VS"].dataset) == 36
        # supplemental qualifiers land in SUPPAE, not in the parent
        assert "AECOMM" not in results["AE"].dataset.columns
        assert len(results["AE"].supp) == 3


def test_deterministic_transforms():
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        _spec, _store, results = _load(tmp)
        dm, ae, vs = results["DM"].dataset, results["AE"].dataset, results["VS"].dataset

        # controlled terminology normalises the raw spellings
        assert list(dm["SEX"]) == ["M", "F", "M", "F", "M", "F"]
        # partial dates keep their real precision and are never padded out
        assert list(dm["BRTHDTC"])[:3] == ["1975-03-12", "1962-11", "1988-07-30"]
        assert list(dm["BRTHDTC"])[5] == ""
        # strip(upcase(x)) in the spec's SAS code is honoured
        assert list(ae["AETERM"])[:2] == ["HEADACHE", "NAUSEA"]
        # --SEQ restarts per subject
        assert list(ae["AESEQ"])[:3] == [1, 2, 1]
        # study day: +1 on or after RFSTDTC, no day zero
        assert list(ae["AESTDY"])[:2] == [6, 19]
        assert vs.loc[0, "VSDY"] == 1                      # measured on RFSTDTC itself
        # --DTC carries the collected time
        assert vs.loc[0, "VSDTC"] == "2024-01-15T09:30"
        # a narrative-only rule is reported, not guessed
        dthfl = next(b for b in results["DM"].blocks if b.variable == "DTHFL")
        assert dthfl.status == "not_built"
        # ...but the dataset is still submission-shaped: the variable is PRESENT and empty,
        # because an SDTM domain carries every variable its spec defines
        assert "DTHFL" in dm.columns
        assert dm["DTHFL"].isna().all()


def test_comparison_finds_planted_differences():
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        _spec, _store, results = _load(tmp)
        comps = compare_study(results, tmp / "vendor")

        ae = comps["AE"]
        assert ae.only_built == 1 and ae.only_vendor == 0     # the dropped COUGH record
        by_var = {d.variable: d for d in ae.diffs}
        assert by_var["AESTDY"].differing == 5                # vendor's off-by-one study day
        assert by_var["AESEV"].differing == 1                 # 'severe' left un-normalised
        assert by_var["AEENDTC"].differing == 0               # end date is compared, not a key
        assert "AEENDTC" not in ae.keys

        dm = comps["DM"]
        assert {d.variable for d in dm.diffs if d.differing} == {"RACE"}
        assert dm.not_built == ["DTHFL"]

        vs = comps["VS"]
        assert vs.total_differences == 0
        assert vs.vars_only_built == ["VSDY"]                 # vendor never delivered it

        assert comps["SUPPAE"].error.startswith("not present")

        # the transposed domain compares like any other — the dropped unit is found
        eg = comps["EG"]
        assert eg.matched == 36 and eg.only_built == 0 and eg.only_vendor == 0
        assert {d.variable for d in eg.diffs if d.differing} == {"EGORRESU"}
        assert next(d for d in eg.diffs if d.variable == "EGORRESU").differing == 12

        # the stacked domain finds the missing disposition record
        ds = comps["DS"]
        assert ds.only_built == 1 and ds.only_vendor == 0


def test_wide_findings_are_transposed():
    """A wide form (one record, several measurement columns) becomes one record per test."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        _spec, _store, results = _load(tmp)
        eg = results["EG"]
        assert eg.ok and eg.prep_step is not None
        assert eg.prep_step.op == "transpose_findings"
        # the spec says raw.eg.*, the study collected it as 'egperf' — found by content
        assert eg.prep_step.params["dataset"] == "egperf"
        # 6 subjects x 2 visits x 3 measurements
        assert len(eg.dataset) == 36
        assert set(eg.dataset["EGTESTCD"]) == {"VR", "PRI", "QRSI"}
        first = eg.dataset[eg.dataset["USUBJID"] == "GS-TEST-001-101-001"]
        assert list(first["EGSEQ"]) == [1, 2, 3, 4, 5, 6]
        # id_vars survive the melt, and the record date still becomes a real --DTC
        assert list(first["VISIT"])[:3] == ["SCREENING"] * 3
        assert list(first["EGDTC"])[:3] == ["2024-01-15"] * 3
        # each measurement keeps its own unit
        units = dict(zip(first["EGTESTCD"], first["EGORRESU"]))
        assert units["VR"] == "beats/min" and units["PRI"] == "msec"


def test_multi_form_domain_is_stacked():
    """A domain with no raw dataset of its own unions the forms that supply its records."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        _spec, _store, results = _load(tmp)
        ds = results["DS"]
        assert ds.ok and ds.prep_step is not None
        assert ds.prep_step.op == "stack"
        assert set(ds.prep_step.params["datasets"]) == {"consent", "enroll", "studcomp"}
        assert len(ds.dataset) == 18                       # 6 subjects x 3 forms
        assert set(ds.dataset["DSCAT"]) == {"PROTOCOL MILESTONE", "DISPOSITION EVENT"}
        one = ds.dataset[ds.dataset["USUBJID"] == "GS-TEST-001-101-001"]
        assert list(one["DSSEQ"]) == [1, 2, 3]
        assert list(one["DSDECOD"]) == ["INFORMED CONSENT OBTAINED", "RANDOMIZED", "COMPLETED"]


def test_prep_can_be_turned_off():
    """prep_mode='off' must actually take effect, and the override must be able to replace it."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        store = RawStore.discover(tmp / "raw")
        from sdtm_builder.build import build_domain

        # With the transpose off, the spec's raw.eg.* resolves to nothing the study collected.
        # The build must SAY so rather than quietly produce an empty or wrong-grain dataset.
        off = build_domain(spec, store, "EG", prep_mode="off")
        assert off.prep_step is None
        assert off.dataset is None
        assert "no raw dataset found for EG" in off.error

        # Naming the real form brings it back at the wide form's own grain, un-melted.
        forced = build_domain(spec, store, "EG", prep_mode="off", base_override="egperf")
        assert forced.prep_step is None
        assert len(forced.dataset) == 12                   # 6 subjects x 2 visits, not melted


def test_build_is_reproducible():
    """The same inputs must give byte-identical output every time. Record ORDER is part of
    that: --SEQ numbers the records, so any non-deterministic ordering silently changes the
    data and would surface as phantom vendor differences."""
    import hashlib
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        digests = []
        for _ in range(3):
            _spec, _store, results = _load(tmp)
            blob = "".join(
                results[d].dataset.to_csv(index=False)
                for d in sorted(results) if results[d].ok)
            digests.append(hashlib.sha256(blob.encode()).hexdigest())
        assert len(set(digests)) == 1, f"build is not reproducible: {digests}"

        # and the melted measurement order follows the spec, not a hash-ordered set
        _spec, _store, results = _load(tmp)
        codes = [m["testcd"] for m in results["EG"].prep_step.params["measures"]]
        assert codes == ["VR", "PRI", "QRSI"]


def test_variable_mapping_can_be_edited_and_is_recorded_as_such():
    """A hand-edited mapping must take effect, survive the automatic repair passes, and be
    marked as a deviation from the spec — an edit that looked spec-derived would quietly
    turn the vendor comparison into a comparison against your own work."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        store = RawStore.discover(tmp / "raw")
        from sdtm_builder.build import build_domain

        plain = build_domain(spec, store, "DM")
        assert list(plain.dataset["RACE"])[2] == "BLACK OR AFRICAN AMERICAN"

        edited = build_domain(spec, store, "DM", edits={"RACE": {
            "mtype": "derived", "recipe": "cond", "note": "vendor uses the short term",
            "args": {"rules": [{"src": {"dataset": "dm", "column": "RACECD"}, "op": "eq",
                                "value": "BLACK OR AFRICAN AMERICAN",
                                "then": {"kind": "text", "text": "BLACK"}}],
                     "else": {"dataset": "dm", "column": "RACECD"}}}})
        assert list(edited.dataset["RACE"])[2] == "BLACK"
        blk = next(b for b in edited.blocks if b.variable == "RACE")
        assert blk.edited and blk.spec_method == "dm.RACECD"
        assert "vendor uses the short term" in blk.edit_note
        assert edited.counts["edited"] == 1
        assert any("not by the spec" in w for w in edited.warnings)


def test_an_edit_survives_the_automatic_passes():
    """--DTC variables are rewritten by the ISO pass. A deliberate mapping must still win."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        store = RawStore.discover(tmp / "raw")
        from sdtm_builder.build import build_domain

        auto = build_domain(spec, store, "VS")
        assert auto.dataset.loc[0, "VSDTC"] == "2024-01-15T09:30"      # ISO pass added the time

        forced = build_domain(spec, store, "VS", edits={"VSDTC": {
            "mtype": "assign", "dataset": "vs", "column": "VSDAT"}})
        assert forced.dataset.loc[0, "VSDTC"] == "2024-01-15"          # the edit won
        assert next(b for b in forced.blocks if b.variable == "VSDTC").edited


def test_dataset_level_dedup():
    """Keeping one record per group is a dataset-level preparation, applied before --SEQ."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        store = RawStore.discover(tmp / "raw")
        from sdtm_builder.build import build_domain

        full = build_domain(spec, store, "VS")
        assert len(full.dataset) == 36
        first = build_domain(spec, store, "VS",
                             dedup={"enabled": True, "keys": ["USUBJID", "VSTESTCD"],
                                    "keep": "first"})
        assert len(first.dataset) == 18
        assert set(first.dataset["VISIT"]) == {"SCREENING"}
        assert list(first.dataset["VSSEQ"])[:3] == [1, 2, 3]           # renumbered after dedup

        last = build_domain(spec, store, "VS",
                            dedup={"enabled": True, "keys": ["USUBJID", "VSTESTCD"],
                                   "keep": "last"})
        assert set(last.dataset["VISIT"]) == {"CYCLE 1 DAY 1"}


def test_full_structure_is_the_default_and_can_be_narrowed():
    """A built domain is submission-shaped: every variable the spec defines, in spec order,
    empty where it could not be populated. Only an explicit spec DROP removes one."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        from sdtm_builder.build import build_domain

        full = build_domain(spec, RawStore.discover(tmp / "raw"), "DM")
        spec_vars = [b.variable for b in full.blocks if b.mtype != "drop" and not b.supp]
        assert list(full.dataset.columns) == spec_vars      # spec order, nothing missing
        assert "COUNTRY" not in full.dataset.columns        # the spec says DROP

        narrow = build_domain(spec, RawStore.discover(tmp / "raw"), "DM", include_unbuilt=False)
        assert "DTHFL" not in narrow.dataset.columns
        assert len(narrow.dataset.columns) < len(full.dataset.columns)


def test_reference_dates_are_built_from_the_spec_sources():
    """A reference date is an extreme across the forms the spec lists — not a copy of whichever
    source happens to be written first."""
    import pandas as pd

    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        raw = tmp / "raw"
        # a spec-shaped case: RFSTDTC drawing on two forms, RFENDTC on one with many visits
        pd.DataFrame({"USUBJID": ["A", "A", "B"], "EXSTDAT": ["2024-03-01", "2024-04-01", "2024-02-02"]}) \
            .to_csv(raw / "exp1.csv", index=False)
        pd.DataFrame({"USUBJID": ["A", "B"], "EXSTDAT": ["2024-01-15", "2024-05-05"]}) \
            .to_csv(raw / "exp2.csv", index=False)

        from sdtm_builder.automap import date_sources_from_spec, extreme_for
        from sdtm_builder.blocks import Block
        store = RawStore.discover(raw)

        b = Block(variable="RFSTDTC", domain="DM",
                  input_variables="raw.exp1.EXSTDAT, raw.exp2.EXSTDAT",
                  mapping_rule="Earliest dose date across all exposure forms.")
        srcs = date_sources_from_spec(b, store)
        assert [(x["dataset"], x["date_col"]) for x in srcs] == [("exp1", "EXSTDAT"), ("exp2", "EXSTDAT")]
        assert extreme_for(b) == "min"

        # the direction follows the spec's own wording, then the variable name
        assert extreme_for(Block(variable="RFENDTC", domain="DM",
                                 mapping_rule="Keep the most recent scheduled visit date.")) == "max"
        assert extreme_for(Block(variable="RFICDTC", domain="DM")) == "min"
        assert extreme_for(Block(variable="RFXENDTC", domain="DM")) == "max"

        # non-date columns the spec also lists are not offered as date sources
        b2 = Block(variable="RFSTDTC", domain="DM",
                   input_variables="raw.exp1.EXSTDAT, raw.exp1.USUBJID")
        assert len(date_sources_from_spec(b2, store)) == 1

        # and an ordinary record-level --DTC is left alone, even when the spec lists several
        # forms: DS.DSSTDTC draws on the forms that ARE its records
        from sdtm_builder.build import build_domain
        spec = load_spec(tmp / "mapping_spec.xlsx")
        ds = build_domain(spec, RawStore.discover(tmp / "raw"), "DS")
        dsstdtc = next(b for b in ds.blocks if b.variable == "DSSTDTC")
        assert dsstdtc.recipe != "date_extreme"
        assert len(ds.dataset) == 18                       # still one record per form per subject
        assert ds.dataset["DSSTDTC"].nunique() > 3         # a date per record, not per subject


def test_date_extreme_joins_on_a_shared_key_and_never_returns_a_silent_blank():
    """The failure this guards against: every dataset is joined on whichever subject key it
    shares with the domain. Aggregating on one key and joining on another looks correct at
    both steps and matches nothing, leaving a column that reads as built and is empty."""
    import pandas as pd
    from sdtm_builder.blocks import Block
    from sdtm_builder.build import BuildContext
    from sdtm_builder.ops import OpError, op_date_extreme

    with tempfile.TemporaryDirectory() as td:
        raw = Path(td)
        pd.DataFrame({"USUBJID": ["S-1", "S-2"], "SUBJID": ["1", "2"]}).to_csv(raw / "dm.csv", index=False)
        pd.DataFrame({"SUBJID": ["1", "1", "2"],                       # keyed on SUBJID only
                      "AEENDAT": ["2024-03-01", "2024-05-01", "2024-02-02"]}).to_csv(raw / "ae.csv", index=False)
        pd.DataFrame({"USUBJID": ["S-1", "S-2"],                       # keyed on USUBJID only
                      "CMENDAT": ["2024-06-09", "2024-01-01"]}).to_csv(raw / "cm.csv", index=False)
        pd.DataFrame({"NOKEY": ["x"], "XDAT": ["2024-01-01"]}).to_csv(raw / "orphan.csv", index=False)

        store = RawStore.discover(raw)
        base = store.get("dm")
        ctx = BuildContext(domain="DM", store=store, base=base,
                           frame=pd.DataFrame(index=base.index), built={}, codelists={})

        b = Block(variable="RFPENDTC", domain="DM", mtype="derived", recipe="date_extreme",
                  args={"func": "max", "group_by": ["USUBJID"], "sources": [
                      {"dataset": "ae", "date_col": "AEENDAT"},
                      {"dataset": "cm", "date_col": "CMENDAT"},
                      {"dataset": "orphan", "date_col": "XDAT"}]})
        got = list(op_date_extreme(ctx, b))
        assert got == ["2024-06-09", "2024-02-02"]     # latest across both usable datasets
        assert "orphan" in b.reason                     # the unusable one is named, not ignored

        earliest = Block(variable="RFSTDTC", domain="DM", mtype="derived", recipe="date_extreme",
                         args={"func": "min", "sources": [
                             {"dataset": "ae", "date_col": "AEENDAT"},
                             {"dataset": "cm", "date_col": "CMENDAT"}]})
        assert list(op_date_extreme(ctx, earliest)) == ["2024-03-01", "2024-01-01"]

        # nothing usable must report, never return an empty column
        bad = Block(variable="X", domain="DM", mtype="derived", recipe="date_extreme",
                    args={"func": "max", "sources": [{"dataset": "orphan", "date_col": "XDAT"}]})
        try:
            op_date_extreme(ctx, bad)
        except OpError as exc:
            assert "no subject key shared" in str(exc)
        else:
            raise AssertionError("an unjoinable source should be reported")


def test_a_mapping_that_produces_nothing_is_reported():
    """A mapping that runs without error and populates nothing is the quietest kind of wrong."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        from sdtm_builder.build import build_domain

        res = build_domain(spec, RawStore.discover(tmp / "raw"), "AE", edits={"AETERM": {
            "mtype": "constant", "value": ""}})            # a constant of nothing
        blk = next(b for b in res.blocks if b.variable == "AETERM")
        assert blk.status == "empty"
        assert "produced no values" in blk.reason
        assert "AETERM" in res.dataset.columns             # still in the structure
        assert res.counts["empty"] == 1


def test_naming_the_date_column_is_enough():
    """Naming the whole-date column must be all a reader has to do: the year/month/day parts
    beside it are found automatically, and they are what keeps a partial date partial."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        from sdtm_builder.build import build_domain

        # BRTHDAT is split into _YYYY/_MM/_DD in the fixture and has no whole-date column,
        # so naming the year part alone must still assemble the full date
        only_year = build_domain(spec, RawStore.discover(tmp / "raw"), "DM", edits={"BRTHDTC": {
            "mtype": "derived", "recipe": "iso_date",
            "args": {"dataset": "dm", "date_col": "BRTHDAT_YYYY"}}})
        got = list(only_year.dataset["BRTHDTC"])
        assert got[0] == "1975-03-12"          # all three parts found and assembled
        assert got[1] == "1962-11"             # the day was never collected — still partial
        assert got[5] == ""                    # nothing collected at all

        # and an explicit override still wins over the automatic match
        forced = build_domain(spec, RawStore.discover(tmp / "raw"), "DM", edits={"BRTHDTC": {
            "mtype": "derived", "recipe": "iso_date",
            "args": {"dataset": "dm", "date_col": "BRTHDAT_YYYY",
                     "y_col": "BRTHDAT_YYYY", "m_col": "", "d_col": ""}}})
        assert list(forced.dataset["BRTHDTC"])[0] == "1975"


def test_the_toc_names_the_studys_domains():
    """The TOC sheet is the spec's own statement of which domains are in this study.
    Active = Y is in; Active = N is out of the default build but stays reviewable."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")

        assert set(spec.toc) == {"DM", "AE", "VS", "EG", "DS", "XX"}
        assert "DM-DATA" not in spec.toc                      # companion rows fold away
        assert spec.is_active("DM") and not spec.is_active("XX")
        assert spec.toc["DM"]["label"] == "Demographics"
        assert spec.toc["AE"]["class"] == "EVENTS"
        # a spec WITHOUT a TOC treats everything as active — nothing changes for it
        assert spec.is_active("NEVER-MENTIONED")

        # deactivating a domain removes it from the default build, and ONLY the default:
        spec.toc["EG"]["active"] = False
        results = build_study(spec, RawStore.discover(tmp / "raw"))
        assert "EG" not in results and "DM" in results
        named = build_study(spec, RawStore.discover(tmp / "raw"), domains=["EG"])
        assert "EG" in named                                   # naming it still builds it


def test_spec_header_below_a_title_row_is_found():
    """Spec workbooks commonly open a sheet with a title banner and put the headings on the
    row below. Assuming row 1 made the entire workbook look empty."""
    import openpyxl

    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        src, dst = tmp / "mapping_spec.xlsx", tmp / "titled_spec.xlsx"
        wb = openpyxl.load_workbook(src)
        for name in ("DM", "AE", "VS", "EG", "DS"):
            wb[name].insert_rows(1)
            wb[name]["A1"] = f"{name}: a title banner above the real headings"
        wb.save(dst)

        plain, titled = load_spec(src), load_spec(dst)
        assert titled.domain_names == plain.domain_names
        assert len(titled.rows("DM")) == len(plain.rows("DM"))
        # row numbers still point at the real spreadsheet rows, now one lower
        assert titled.rows("DM")[0].row_number == plain.rows("DM")[0].row_number + 1


def test_a_blank_template_says_so():
    import openpyxl

    with tempfile.TemporaryDirectory() as td:
        blank = Path(td) / "template.xlsx"
        wb = openpyxl.Workbook()
        for i, name in enumerate(("DM", "AE")):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = name
            ws["A1"] = f"{name}: Subject data"
            for col, head in enumerate(["Variable", "Label", "Type", "Origin"], start=1):
                ws.cell(row=2, column=col, value=head)
        wb.save(blank)
        try:
            load_spec(blank)
        except ValueError as exc:
            assert "blank specification template" in str(exc), exc
        else:
            raise AssertionError("a blank template should be reported as such")


def test_sas_code_in_the_spec_is_compiled():
    """The Implemented SAS Code column carries real derivations. Reading only Input Variables
    would turn `scan(usubjid, -2, '-')` into a copy of USUBJID — a confident wrong value."""
    from sdtm_builder.sasexpr import Unresolved, to_block_args

    def resolve(name):
        if name.upper() == "USUBJID":
            return {"kind": "var", "var": "USUBJID"}
        raise Unresolved(name)

    got = to_block_args("scan(usubjid, -2, '-')", resolve)
    assert got["recipe"] == "pipeline"
    step = got["args"]["steps"][0]["args"]
    assert step["fn"] == "scan" and step["word"] == "-2" and step["delim"] == "-"

    joined = to_block_args("scan(usubjid,-2,'-')||'-'||scan(usubjid,-1,'-')", resolve)
    assert len(joined["args"]["steps"]) == 3            # two scans and a concatenation

    # anything outside the supported grammar is refused, never approximated
    assert to_block_args("put(datepart(x), yymmdd10.)", resolve) is None
    assert to_block_args("if a then b; else c;", resolve) is None


def test_name_matching_fills_gaps_the_spec_leaves_and_labels_them():
    """Coverage the spec does not provide, recovered deterministically — and never disguised
    as a spec-derived mapping."""
    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        spec = load_spec(tmp / "mapping_spec.xlsx")
        from sdtm_builder.build import build_domain

        on = build_domain(spec, RawStore.discover(tmp / "raw"), "DM")
        ethnic = next(b for b in on.blocks if b.variable == "ETHNIC")
        assert ethnic.status == "built"
        assert ethnic.method_source == "name_match"     # labelled as the guess it is
        assert (ethnic.dataset, ethnic.column) == ("dm", "ETHNICCD")
        assert 70 <= ethnic.confidence < 100
        assert on.counts["name_matched"] == 1
        assert any("not evidence" in w for w in on.warnings)

        off = build_domain(spec, RawStore.discover(tmp / "raw"), "DM", name_match_threshold=0)
        assert next(b for b in off.blocks if b.variable == "ETHNIC").status == "not_built"
        assert off.counts["name_matched"] == 0

        # a variable with no similar raw column is never invented, at any threshold
        for thr in (0, 45, 70):
            r = build_domain(spec, RawStore.discover(tmp / "raw"), "DM",
                             name_match_threshold=thr)
            assert next(b for b in r.blocks if b.variable == "DTHFL").status == "not_built"

        # a spec-derived mapping keeps its provenance
        assert next(b for b in on.blocks if b.variable == "RACE").method_source == "spec"


def test_best_listed_source_is_preferred_over_the_first():
    """When the spec lists several sources, the one that exists and fits the variable wins."""
    import pandas as pd
    from sdtm_builder.automap import best_listed_source

    with tempfile.TemporaryDirectory() as td:
        tmp = _fixture(Path(td))
        raw = tmp / "raw"
        pd.DataFrame({"USUBJID": ["A"], "AESEV_STD": ["MILD"], "AETOXGR_STD": ["1"]}) \
            .to_csv(raw / "aegrade.csv", index=False)
        store = RawStore.discover(raw)
        iv = "raw.aegrade.AESEV_STD, raw.aegrade.AETOXGR_STD"
        assert best_listed_source("AETOXGR", iv, store) == ("aegrade", "AETOXGR_STD")
        assert best_listed_source("AESEV", iv, store) == ("aegrade", "AESEV_STD")


def test_numeric_and_blank_equivalence():
    """1, 1.0 and '1.00' are the same value; '' and NA are the same absence."""
    import pandas as pd
    from sdtm_builder.compare import compare_domain
    built = pd.DataFrame({"USUBJID": ["A", "B"], "XXSEQ": [1, 2],
                          "XXORRES": [1.0, 2.50], "XXCOM": ["", None]})
    vendor = pd.DataFrame({"USUBJID": ["A", "B"], "XXSEQ": ["1", "2"],
                           "XXORRES": ["1.00", "2.5"], "XXCOM": [None, ""]})
    c = compare_domain("XX", built, vendor, keys=["USUBJID"])
    assert c.total_differences == 0, [(d.variable, d.examples) for d in c.diffs if d.differing]


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"PASS  {name}")
            except AssertionError as exc:
                failures += 1
                print(f"FAIL  {name}: {exc}")
    print(f"\n{'all tests passed' if not failures else f'{failures} test(s) failed'}")
    raise SystemExit(1 if failures else 0)
